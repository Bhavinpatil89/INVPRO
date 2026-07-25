from flask import Blueprint, render_template, request, jsonify, session, send_file
from ..core.auth import login_required
from ..core.db import get_all_bills, get_bill_items, execute_query
from ..core.charts import generate_sales_chart, generate_profit_chart, generate_tax_chart
from datetime import datetime, timedelta
import io
from openpyxl import Workbook

sales_bp = Blueprint('sales', __name__)

def get_date_range(time_filter):
    """Utility to get start date based on filter."""
    today = datetime.now().date()
    if time_filter == 'today':
        return today
    elif time_filter == 'week':
        return today - timedelta(days=7)
    elif time_filter == 'month':
        return today - timedelta(days=30)
    elif time_filter == 'year':
        return today - timedelta(days=365)
    elif time_filter == 'financial_year':
        # Indian Financial Year: April 1 to March 31
        current_year = today.year
        if today.month < 4:
            return datetime(current_year - 1, 4, 1).date()
        else:
            return datetime(current_year, 4, 1).date()
    return None

@sales_bp.route('/history')
@login_required
def history():
    return render_template('sales_history.html')

@sales_bp.route('/api/history')
@login_required
def get_sales_history_api():
    """API to get all bills with time filters."""
    role = session.get('role')
    store_id = session.get('store_id')
    chain_id = session.get('chain_id')

    if role == 'business_admin':
         # HQ View: Show all bills for the chain
         scope_clause = "b.store_id IN (SELECT id FROM stores WHERE chain_id = ?)"
         scope_params = [chain_id]
    elif store_id:
         # Branch View: Locked to store
         scope_clause = "b.store_id = ?" 
         scope_params = [store_id]
    else:
        return jsonify([])
    
    time_filter = request.args.get('filter', 'all')
    start_date = get_date_range(time_filter)
    
    query = f"""
        SELECT b.*, u.username as creator_name, COUNT(bi.id) as item_count 
        FROM bills b 
        LEFT JOIN bill_items bi ON b.id = bi.bill_id 
        LEFT JOIN users u ON b.user_id = u.id
        WHERE {scope_clause}
    """
    params = list(scope_params)
    
    if start_date:
        query += " AND date(b.date) >= date(?)"
        params.append(start_date)
        
    query += " GROUP BY b.id, u.username ORDER BY b.date DESC"

    bills = execute_query(query, tuple(params), fetch_all=True)
    return jsonify([dict(b) for b in bills])

from ..services.context import UserContext
from ..services.reports import ReportService

@sales_bp.route('/api/analytics')
@login_required
def get_analytics():
    """Get sales analytics with time filter."""
    context = UserContext(
        user_id=session.get('user_id'),
        username=session.get('username'),
        role=session.get('role'),
        chain_id=session.get('chain_id'),
        store_id=session.get('store_id')
    )
    
    report_service = ReportService(context)
    time_filter = request.args.get('filter', 'all')
    store_filter_id = request.args.get('branch')
    
    if context.is_business_admin():
        # Allow HQ to filter by specific store for "Focus Mode"
        data = report_service.get_enterprise_dashboard_stats(time_filter, store_id=store_filter_id)
    else:
        # Branch users locked to their store
        data = report_service.get_store_dashboard_stats(time_filter)
        
    return jsonify(data)

@sales_bp.route('/api/charts')
@login_required
def get_charts_api():
    """Serve JSON data for Chart.js frontend."""
    context = UserContext(
        user_id=session.get('user_id'),
        username=session.get('username'),
        role=session.get('role'),
        chain_id=session.get('chain_id'),
        store_id=session.get('store_id')
    )
    
    report_service = ReportService(context)
    time_filter = request.args.get('filter', 'month')
    store_filter_id = request.args.get('branch')
    
    return jsonify(report_service.get_chart_data(time_filter, store_id=store_filter_id))

@sales_bp.route('/api/comparison')
@login_required
def get_comparison_api():
    """Get comparison data for selected branches."""
    context = UserContext(
        user_id=session.get('user_id'),
        username=session.get('username'),
        role=session.get('role'),
        chain_id=session.get('chain_id'),
        store_id=session.get('store_id')
    )
    if not context.is_business_admin():
        return jsonify({'error': 'Unauthorized'}), 403

    report_service = ReportService(context)
    branch_ids_param = request.args.get('branches')
    time_filter = request.args.get('time_filter', 'month')
    
    branch_ids = None
    if branch_ids_param:
        try:
            branch_ids = [int(bid) for bid in branch_ids_param.split(',')]
        except ValueError:
            pass # Fallback to top 10

    return jsonify(report_service.get_branch_comparison(branch_ids, time_filter))

@sales_bp.route('/export')
@login_required
def export_sales():
    """Export sales history to Excel."""
    role = session.get('role')
    user_chain_id = session.get('chain_id')
    
    # Resolve target store
    target_store_id = request.args.get('store_id') or session.get('store_id')
    
    # Security Check
    if not target_store_id and role != 'business_admin':
        return "Unauthorized", 403
        
    if target_store_id:
        target_store_id = int(target_store_id)
        # Verify ownership
        from ..core.db import get_all_stores_by_chain
        stores = get_all_stores_by_chain(user_chain_id)
        if not any(s['id'] == target_store_id for s in stores) and role == 'business_admin':
             return "Unauthorized access to store", 403
        elif role != 'business_admin' and target_store_id != session.get('store_id'):
             return "Unauthorized access to branch data", 403

    time_filter = request.args.get('filter', 'all')
    start_date = get_date_range(time_filter)
    
    query = "SELECT * FROM bills WHERE 1=1"
    params = []
    
    if target_store_id:
        query += " AND store_id = ?"
        params.append(target_store_id)
    else:
        # HQ combined export
        query += " AND store_id IN (SELECT id FROM stores WHERE chain_id = ?)"
        params.append(user_chain_id)
        
    if start_date:
        query += " AND DATE(date) >= DATE(?)"
        params.append(start_date)
        
    query += " ORDER BY date DESC"
    bills = execute_query(query, tuple(params), fetch_all=True)
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    headers = ['Bill ID', 'Bill Number', 'Date', 'Subtotal', 'Tax', 'Total Amount', 'Profit']
    ws.append(headers)
    
    for b in bills:
        ws.append([b['id'], b['bill_number'], b['date'], b['subtotal_amount'], b['tax_amount'], b['total_amount'], b['total_profit']])
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"sales_report_{time_filter}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)

@sales_bp.route('/reports')
@login_required
def report_center():
    """Render the Reports Dashboard."""
    return render_template('reports.html')

@sales_bp.route('/export/inventory')
@login_required
def export_inventory_valuation():
    """Export current inventory value."""
    role = session.get('role')
    user_chain_id = session.get('chain_id')
    target_store_id = request.args.get('store_id') or session.get('store_id')
    
    # Security
    if target_store_id:
        target_store_id = int(target_store_id)
        from ..core.db import get_all_stores_by_chain
        stores = get_all_stores_by_chain(user_chain_id)
        if not any(s['id'] == target_store_id for s in stores) and role == 'business_admin':
             return "Unauthorized", 403

    query = """
        SELECT p.*, c.name as category_name 
        FROM products p 
        JOIN categories c ON p.category_id = c.id 
        WHERE p.quantity > 0
    """
    params = []
    
    if target_store_id:
        query += " AND p.store_id = ?"
        params.append(target_store_id)
    else:
        query += " AND p.store_id IN (SELECT id FROM stores WHERE chain_id = ?)"
        params.append(user_chain_id)
        
    query += " ORDER BY p.name"
    products = execute_query(query, tuple(params), fetch_all=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Valuation"
    
    headers = ['Category', 'Product Name', 'Quantity', 'Cost Price', 'Selling Price', 'Total Cost', 'Total Sales', 'Profit']
    ws.append(headers)
    
    for p in products:
        cost_val = p['cost_price'] * p['quantity']
        sales_val = p['selling_price'] * p['quantity']
        ws.append([p['category_name'], p['name'], p['quantity'], p['cost_price'], p['selling_price'], cost_val, sales_val, (sales_val - cost_val)])
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name="inventory_valuation.xlsx")

@sales_bp.route('/export/products')
@login_required
def export_product_performance():
    """Export product sales performance."""
    store_id = session.get('store_id')
    role = session.get('role')
    chain_id = session.get('chain_id')
    
    scope = 'chain' if role == 'business_admin' else 'store'
    scope_id = chain_id if role == 'business_admin' else store_id
    
    # Import helper from db here to avoid circular import issues if placed at top level with app context
    from ..core.db import get_top_selling_products
    
    # Get top 500 products (effectively all relevant ones)
    data = get_top_selling_products(scope, scope_id, limit=500)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Performance"
    
    headers = ['Product Name', 'Units Sold', 'Total Revenue', 'Total Profit', 'Margin %']
    ws.append(headers)
    
    for row in data:
        rev = row['total_revenue']
        prof = row['total_profit']
        margin = (prof / rev * 100) if rev else 0
        
        ws.append([
            row['product_name'],
            row['total_qty'],
            rev,
            prof,
            f"{margin:.2f}%"
        ])
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"product_performance_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)

@sales_bp.route('/export/tax')
@login_required
def export_tax_audit():
    """Export detailed Tax Audit report."""
    role = session.get('role')
    user_chain_id = session.get('chain_id')
    target_store_id = request.args.get('store_id') or session.get('store_id')
    time_filter = request.args.get('filter', 'month')
    start_date = get_date_range(time_filter)

    # Security Check
    if target_store_id:
        target_store_id = int(target_store_id)
        from ..core.db import get_all_stores_by_chain
        stores = get_all_stores_by_chain(user_chain_id)
        if not any(s['id'] == target_store_id for s in stores) and role == 'business_admin':
             return "Unauthorized", 403

    query = "SELECT bill_number, date, subtotal_amount, tax_amount, total_amount FROM bills WHERE 1=1"
    params = []
    
    if target_store_id:
        query += " AND store_id = ?"
        params.append(target_store_id)
    else:
        query += " AND store_id IN (SELECT id FROM stores WHERE chain_id = ?)"
        params.append(user_chain_id)
        
    if start_date:
        query += " AND DATE(date) >= DATE(?)"
        params.append(start_date)
        
    query += " ORDER BY date DESC"
    bills = execute_query(query, tuple(params), fetch_all=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Tax Audit Report"
    
    headers = ['Bill Number', 'Date', 'Taxable Amount', 'Tax Collected', 'Total Amount']
    ws.append(headers)
    
    total_taxable = 0
    total_tax = 0
    
    for b in bills:
        ws.append([b['bill_number'], b['date'], b['subtotal_amount'], b['tax_amount'], b['total_amount']])
        total_taxable += b['subtotal_amount']
        total_tax += b['tax_amount']
        
    ws.append([])
    ws.append(['TOTALS:', '', total_taxable, total_tax, (total_taxable + total_tax)])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"tax_audit_{time_filter}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)

@sales_bp.route('/api/bill/<int:bill_id>')
@login_required
def get_bill_details(bill_id):
    """API to get bill items for modal, enforcing context security."""
    store_id = session.get('store_id')
    chain_id = session.get('chain_id')
    role = session.get('role')
    
    if role == 'business_admin' and chain_id:
        # Check if bill belongs to any store in owner's chain
        bill = execute_query("""
            SELECT b.id FROM bills b 
            JOIN stores s ON b.store_id = s.id 
            WHERE b.id = ? AND s.chain_id = ?
        """, (bill_id, chain_id), fetch_one=True)
        if not bill:
            return jsonify({'error': 'Unauthorized'}), 403
            
        # Get items (we need a way to bypass get_bill_items's store_id check or fetch carefully)
        # Re-using raw query for full chain visibility
        items = execute_query("SELECT * FROM bill_items WHERE bill_id = ?", (bill_id,), fetch_all=True)
    else:
        # Strict store check
        items = get_bill_items(bill_id, store_id)
        
    return jsonify([dict(item) for item in items])
