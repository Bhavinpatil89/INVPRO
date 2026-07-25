from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from ..core.auth import login_required
from ..services.auth import AuthService
from ..services.inventory import InventoryService
from openpyxl import Workbook, load_workbook
from io import BytesIO



# ==================== CONFIGURATION ====================
inventory_bp = Blueprint('inventory', __name__)
auth_service = AuthService()

def get_inventory_service():
    """Resolve InventoryService with active context."""
    user_id = session.get('user_id')
    context = auth_service.build_user_context(
        user_id, 
        session_chain_id=session.get('chain_id'), 
        session_store_id=session.get('store_id')
    )
    return InventoryService(context)



# ==================== VIEW ROUTES (CSR SHELLS) ====================



@inventory_bp.route('/global')
@login_required
def global_inventory():
    """Master Product List (HQ Admin Focus)."""
    return render_template('inventory/global.html')

@inventory_bp.route('/categories')
@login_required
def categories():
    """Product Categories (HQ Admin Focus)."""
    service = get_inventory_service()
    if not service.context.can_manage_inventory_structure():
        return redirect(url_for('main.dashboard'))
    return render_template('categories.html')

@inventory_bp.route('/products')
@login_required
def products():
    """Shop Inventory Control."""
    return render_template('products.html')

@inventory_bp.route('/import-view')
@login_required
def import_registry_view():
    """Add Multiple Items from Master List to Shop."""
    try:
        service = get_inventory_service()
        if service.context.is_business_admin() and not service.context.store_id:
            flash("The Master List is global. Select a Shop to add products.", 'info')
            return redirect(url_for('inventory.global_inventory'))
        
        return render_template('inventory/import.html')
    except Exception as e:
        flash(f"Error: {str(e)}", 'error')
        return redirect(url_for('main.dashboard'))



@inventory_bp.route('/api/categories')
@login_required
def get_categories_api():
    """Fetch categories with product and tax counts."""
    try:
        chain_id = session.get('chain_id')
        
        # Get categories with product and tax counts
        from ..core.db import execute_query
        categories = execute_query("""
            SELECT c.*,
                   (SELECT COUNT(*) FROM global_products WHERE category_id = c.id) as product_count,
                   (SELECT COUNT(DISTINCT tp.id) 
                    FROM tax_policies tp 
                    JOIN tax_policy_categories tpc ON tp.id = tpc.tax_id 
                    WHERE tpc.category_id = c.id) as tax_count
            FROM categories c
            WHERE c.chain_id = ?
            ORDER BY c.name
        """, (chain_id,), fetch_all=True)
        
        return jsonify([dict(c) for c in categories])
    except Exception as e:
        return jsonify({'error': str(e)}), 403

def _handle_action(func, success_msg, redirect_to='inventory.products'):
    """Helper to consolidate try-except-flash pattern. Supports JSON and Form."""
    try:
        func()
        if request.is_json:
            return jsonify({'success': True, 'message': success_msg})
        flash(success_msg, 'success')
    except Exception as e:
        if request.is_json:
            return jsonify({'error': str(e)}), 400
        flash(f"Error: {str(e)}", 'error')
    
    return redirect(url_for(redirect_to))

@inventory_bp.route('/categories/add', methods=['POST'])
@login_required
def add_category():
    def action():
        service = get_inventory_service()
        if not service.context.can_manage_inventory_structure(): raise PermissionError("Insufficient authority.")
        name, tax = request.form.get('name', '').strip(), request.form.get('tax', '').strip()
        if not name: raise ValueError("Name required.")
        service.add_category(name)
        if tax:
            from ..core.db import execute_query
            res = execute_query("SELECT id FROM categories WHERE name = ? AND chain_id = ? ORDER BY id DESC", (name, service.context.chain_id), fetch_one=True)
            if res:
                from ..services.configuration import ConfigurationService
                ConfigurationService(service.context).set_chain_setting(f"tax_category_{res['id']}", tax)
    return _handle_action(action, "Category created.", 'inventory.categories')

@inventory_bp.route('/categories/edit/<int:category_id>', methods=['POST'])
@login_required
def edit_category(category_id):
    def action():
        service = get_inventory_service()
        if not service.context.can_manage_inventory_structure(): raise PermissionError("Insufficient authority.")
        name, tax = request.form.get('name', '').strip(), request.form.get('tax', '').strip()
        if not name: raise ValueError("Name required.")
        service.update_category(category_id, name)
        from ..services.configuration import ConfigurationService
        config = ConfigurationService(service.context)
        if service.context.is_business_admin(): config.set_chain_setting(f"tax_category_{category_id}", tax)
        else: config.set_store_setting(f"tax_category_{category_id}", tax)
    return _handle_action(action, "Category updated.", 'inventory.categories')

@inventory_bp.route('/categories/delete/<int:category_id>', methods=['POST'])
@login_required
def delete_category_route(category_id):
    def action():
        service = get_inventory_service()
        if not service.context.can_manage_inventory_structure(): raise PermissionError("Insufficient authority.")
        service.delete_category(category_id)
    return _handle_action(action, "Category deleted.", 'inventory.categories')

# ==================== PRODUCT MANAGEMENT ====================

@inventory_bp.route('/products/add', methods=['POST'])
@login_required
def add_product():
    def action():
        service = get_inventory_service()
        if not service.context.is_business_admin(): raise PermissionError("Direct creation restricted to HQ. Please use 'Import from HQ' instead.")
        name, tax = request.form.get('name', '').strip(), request.form.get('tax', '').strip()
        if not name: raise ValueError("Name required.")
        service.add_product({
            'name': name, 'cost_price': float(request.form.get('cost_price', 0)),
            'selling_price': float(request.form.get('selling_price', 0)),
            'quantity': int(request.form.get('quantity', 0)),
            'category_id': int(request.form.get('category_id'))
        })
        if tax:
            from ..core.db import execute_query
            res = execute_query("SELECT id FROM products WHERE name = ? AND store_id = ? ORDER BY id DESC", (name, service.context.store_id), fetch_one=True)
            if res:
                from ..services.configuration import ConfigurationService
                ConfigurationService(service.context).set_item_tax_override(res['id'], float(tax))
    return _handle_action(action, "Product registered.")

@inventory_bp.route('/products/edit/<int:product_id>', methods=['POST', 'PUT'])
@login_required
def edit_product(product_id):
    def action():
        service = get_inventory_service()
        if service.context.is_branch_staff(): raise PermissionError("Read Only.")
        
        data = request.get_json() if request.is_json else request.form
        tax = data.get('tax', '').strip() if hasattr(data, 'get') else ''
        
        if not service.context.can_manage_inventory_structure():
            curr = service.get_product_by_id(product_id)
            if not curr: raise ValueError("Product not found.")
            service.update_product(product_id, {
                'name': curr['name'], 'category_id': curr['category_id'], 'cost_price': curr['cost_price'],
                'selling_price': float(data.get('selling_price', 0)) or curr['selling_price'],
                'quantity': int(data.get('quantity', 0))
            })
        else:
            service.update_product(product_id, {
                'name': data.get('name', '').strip(),
                'cost_price': float(data.get('cost_price', 0)),
                'selling_price': float(data.get('selling_price', 0)),
                'quantity': int(data.get('quantity', 0)),
                'category_id': int(data.get('category_id'))
            })
            from ..services.configuration import ConfigurationService
            config = ConfigurationService(service.context)
            if service.context.is_business_admin(): config.set_chain_setting(f"tax_product_{product_id}", tax)
            else: config.set_store_setting(f"tax_product_{product_id}", tax)
    return _handle_action(action, "Product updated.")

@inventory_bp.route('/products/delete/<int:product_id>', methods=['POST', 'DELETE'])
@login_required
def delete_product_route(product_id):
    def action():
        service = get_inventory_service()
        if not (service.context.is_business_admin() or service.context.is_branch_admin()): raise PermissionError("Authority Restricted.")
        service.delete_product(product_id)
    return _handle_action(action, "Product removed.")

# ==================== EXCEL ENGINE ====================

@inventory_bp.route('/products/import', methods=['POST'])
@login_required
def import_products():
    """Bulk upload of products via Excel."""
    try:
        service = get_inventory_service()
        if not service.context.can_manage_inventory_structure(): raise PermissionError("Restricted Operation.")
        if 'file' not in request.files:
            flash('No file.', 'error')
            return redirect(url_for('inventory.products'))
        
        file = request.files['file']
        wb = load_workbook(file)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        required = ['Product Name', 'Category', 'Cost Price', 'Selling Price', 'Quantity']
        
        if not all(col in headers for col in required):
            flash('Invalid Template.', 'error')
            return redirect(url_for('inventory.products'))
            
        col_indices = {col: headers.index(col) for col in required}
        cats = service.get_chain_categories()
        cat_map = {c['name']: c['id'] for c in cats}
        
        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                name, cat_name = str(row[col_indices['Product Name']]).strip(), str(row[col_indices['Category']]).strip()
                cost, sell, qty = float(row[col_indices['Cost Price']]), float(row[col_indices['Selling Price']]), int(row[col_indices['Quantity']])
                if cat_name in cat_map:
                    service.add_product({'name': name, 'cost_price': cost, 'selling_price': sell, 'quantity': qty, 'category_id': cat_map[cat_name]})
                    imported += 1
            except: continue
            
        flash(f'Added {imported} items successfully.', 'success')
    except Exception as e: flash(f'Import error: {str(e)}', 'error')
    return redirect(url_for('inventory.products'))

@inventory_bp.route('/products/export')
@login_required
def export_products():
    """Export Node Registry to Excel."""
    try:
        service = get_inventory_service()
        products = service.get_store_products()
        wb = Workbook()
        ws = wb.active
        ws.append(['ID', 'Name', 'Category', 'Cost', 'Sell', 'Qty', 'Profit'])
        for p in products:
            ws.append([p['id'], p['name'], p.get('category_name'), p['cost_price'], p['selling_price'], p['quantity'], p.get('profit')])
            
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='inventory.xlsx')
    except Exception as e:
        flash(str(e), 'error')
        return redirect(url_for('inventory.products'))

# ==================== GLOBAL REGISTRY API ====================

@inventory_bp.route('/api/global-products', methods=['GET'])
@login_required
def get_global_products():
    """Fetch all products in the Enterprise HQ Registry."""
    try:
        from ..core.db import execute_query
        chain_id = session.get('chain_id')
        if not chain_id: return jsonify([]), 403
        
        products = execute_query("""
            SELECT gp.*, c.name as category_name
            FROM global_products gp
            LEFT JOIN categories c ON gp.category_id = c.id
            WHERE gp.chain_id = ?
            ORDER BY gp.name
        """, (chain_id,), fetch_all=True)
        
        return jsonify([dict(p) for p in products])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@inventory_bp.route('/api/global-products', methods=['POST'])
@login_required
def create_global_product():
    """Register a new product in the Global HQ Registry."""
    try:
        chain_id = session.get('chain_id')
        if session.get('role') != 'business_admin': return jsonify({'error': 'Unauthorized'}), 403
        
        data = request.get_json()
        from ..core.db import execute_query
        product_id = execute_query("""
            INSERT INTO global_products (name, default_cost_price, default_selling_price, category_id, chain_id)
            VALUES (?, ?, ?, ?, ?)
        """, (data['name'], data['default_cost_price'], data['default_selling_price'], data['category_id'], chain_id))
        
        return jsonify({'id': product_id, 'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@inventory_bp.route('/api/global-products/<int:product_id>', methods=['PUT'])
@login_required
def update_global_product(product_id):
    """Update a product in the Global HQ Registry."""
    try:
        chain_id = session.get('chain_id')
        if session.get('role') != 'business_admin': return jsonify({'error': 'Unauthorized'}), 403
        
        data = request.get_json()
        from ..core.db import execute_query
        
        execute_query("""
            UPDATE global_products 
            SET name = ?, default_cost_price = ?, default_selling_price = ?, category_id = ?
            WHERE id = ? AND chain_id = ?
        """, (data['name'], data['default_cost_price'], data['default_selling_price'], data['category_id'], product_id, chain_id))
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@inventory_bp.route('/api/global-products/<int:product_id>', methods=['DELETE'])
@login_required
def delete_global_product(product_id):
    """Remove a product from the Global HQ Registry."""
    try:
        chain_id = session.get('chain_id')
        if session.get('role') != 'business_admin': return jsonify({'error': 'Unauthorized'}), 403
        
        from ..core.db import execute_query
        execute_query("DELETE FROM global_products WHERE id = ? AND chain_id = ?", (product_id, chain_id))
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@inventory_bp.route('/api/global-products/export')
@login_required
def export_global_products():
    """Export HQ Registry to Excel."""
    try:
        from ..core.db import execute_query
        chain_id = session.get('chain_id')
        if session.get('role') != 'business_admin': return "Unauthorized", 403
        
        products = execute_query("""
            SELECT gp.*, c.name as category_name
            FROM global_products gp
            LEFT JOIN categories c ON gp.category_id = c.id
            WHERE gp.chain_id = ?
            ORDER BY gp.name
        """, (chain_id,), fetch_all=True)
        
        wb = Workbook()
        ws = wb.active
        ws.append(['ID', 'Name', 'Category', 'Default Cost', 'Default Sell'])
        for p in products:
            ws.append([p['id'], p['name'], p['category_name'], p['default_cost_price'], p['default_selling_price']])
            
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='hq_master_list.xlsx')
    except Exception as e:
        return str(e), 500

# ==================== IMPORT PROTOCOLS ====================

@inventory_bp.route('/api/import-products', methods=['POST'])
@login_required
def import_products_api():
    """Deploy assets from Global Registry to a dynamic Node."""
    try:
        store_id = session.get('store_id')
        if not store_id: return jsonify({'error': 'No Node Context.'}), 403
        
        data = request.get_json()
        products = data.get('products', [])
        if not products: return jsonify({'error': 'Empty Payload.'}), 400
        
        from ..core.db import get_db_connection
        imported_count = 0
        with get_db_connection() as conn:
            cursor = conn.cursor()
            for product in products:
                cursor.execute("SELECT id FROM products WHERE global_product_id = ? AND store_id = ?", (product['global_product_id'], store_id))
                if cursor.fetchone(): continue
                
                cursor.execute("""
                    INSERT INTO products (global_product_id, name, cost_price, selling_price, quantity, category_id, store_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (product['global_product_id'], product['name'], product['cost_price'], product['selling_price'], product.get('quantity', 0), product['category_id'], store_id))
                imported_count += 1
        
        return jsonify({'success': True, 'imported': imported_count})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@inventory_bp.route('/api/products', methods=['GET'])
@login_required
def get_branch_products_api():
    """Fetch Node products with context resolution."""
    try:
        service = get_inventory_service()
        
        # External Context Injection: Allow HQ to query any node in their scope
        requested_store_id = request.args.get('store_id')
        if requested_store_id and service.context.is_business_admin():
            from ..services.context import UserContext
            from ..services.auth import AuthService
            auth = AuthService()
            # Build a temporary context for the requested store
            ctx = auth.build_user_context(session.get('user_id'), session_chain_id=session.get('chain_id'), session_store_id=int(requested_store_id))
            service = InventoryService(ctx)
            
        if not service.store_id: return jsonify({'error': 'No Node Context.'}), 403
        return jsonify(service.get_store_products())
    except Exception as e:
        return jsonify({'error': str(e)}), 500
